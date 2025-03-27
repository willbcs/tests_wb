import openpyxl  # biblioteca para abrir excel
from selenium import webdriver  # biblioteca para acesso a internet
from selenium.webdriver.common.by import By  # Permite encontrar elementos dentro da página
from time import sleep
from PySimpleGUI import PySimpleGUI as sg
#from selenium.webdriver.support.ui import WebDriverWait
#from selenium.webdriver.support import expected_conditions as EC

sg.theme('Reddit')
def resultados_tela(dados):
    layout = [
        [sg.Table(values=dados, headings=["Nome", "CPF", "Status"], auto_size_columns = True, num_rows=30, justification="left")],
        [sg.Button('Fechar')]
    ]
    janela = sg.Window('Validador de CPF', layout)

    while True:
        eventos, valores = janela.read()
        if eventos == sg.WINDOW_CLOSED or eventos == 'Fechar':
            break
    
    janela.close()

# Abrir planilha e indicar a aba
planilha_CPF = openpyxl.load_workbook('C:/Users/Myrath/Desktop/Projetos_VSCode/.vscode/.python/projeto_verifica_CPF/cpf_clientes.xlsx')
pagina_CPF = planilha_CPF['Sheet1']

# Acessar o site e entrar com as informações de CPF a serem validadas
driver = webdriver.Chrome()
driver.get('https://validadordecpf.clevert.com.br/v-cpf.php')
sleep(4)

validacao_CPF = openpyxl.load_workbook('C:/Users/Myrath/Desktop/Projetos_VSCode/.vscode/.python/projeto_verifica_CPF/validacao_CPF.xlsx')
pagina_validacao = validacao_CPF['Sheet1']

resultados = []

# Ler as linhas a partir da linha 02 e processar o Nome e CPF
for linha in pagina_CPF.iter_rows(min_row=2, values_only=True):
    nome, cpf = linha

    # Localiza o campo para digitar CPF e limpa antes de enviar o número
    campo_consulta = driver.find_element(By.XPATH, "//input[@id='cpf']")
    campo_consulta.clear()
    campo_consulta.click()
    campo_consulta.send_keys(cpf)

    # Clica no botão de consulta
    botao_consulta = driver.find_element(By.XPATH, "//button[@id='gerar']")
    botao_consulta.click()

    # Espera o resultado da validação ser atualizado
    sleep(2)  # Espera mais tempo para garantir que a resposta foi carregada

    # Verifica se o resultado do CPF é válido ou inválido
    #validacao = driver.find_element(By.XPATH, "//div[@role='alert']")
    try:
        validacao = driver.find_element(By.XPATH, "//div[@id='resposta1']")
        
    except:
        validacao = driver.find_element(By.XPATH, "//div[@id='resposta2']")

    #validacao = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@role='alert']")))
    sleep(1) 
    resultado_texto = validacao.text

    if 'válido' in resultado_texto:
        status_validacao = 'CPF Válido!'

    else:
        status_validacao = 'CPF Inválido!'
    
    resultados.append([nome, cpf, status_validacao])

    pagina_validacao.append([nome, cpf, status_validacao])
    
validacao_CPF.save('C:/Users/Myrath/Desktop/Projetos_VSCode/.vscode/.python/projeto_verifica_CPF/validacao_CPF.xlsx')
# Fecha o navegador após acabar a rotina
driver.quit()

resultados_tela(resultados)