import openpyxl # biblioteca para abrir excel
from selenium import webdriver #biblioteca para acesso a internet
from selenium.webdriver.common.by import By #Permite encontrar elementos dentro da página
from time import sleep
# from selenium.webdriver.support.ui import WebDriverWait #Para esperas explícitas
# from selenium.webdriver.support import expected_conditions as EC #Condições esperadas para esperas

#Abrir planilha e a aba
planilha_CPF = openpyxl.load_workbook('cpf_clientes.xlsx')
pagina_CPF = planilha_CPF['Sheet1']

#acessar o Site e entrar com as informações de CPF a serem validadas | 
# formato do XPATH //tag[@atributo='valor']
driver = webdriver.Chrome()
driver.get('https://www.geradorcpf.com/validar-cpf.htm')

#Ler as linhas, a partir da 02 e quando preenchidas, puxar Nome e CPF
for linha in pagina_CPF.iter_rows(min_row=2, values_only=True):
    nome, cpf = linha   

# Espera até que o campo de CPF esteja visível e interativo
    # campo_consulta = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='cpf']")))
# Poderia ser assim também...mas o tempo seria cronometrado e teria que existir o "from time import sleep"
    sleep(4)
    #localiza o campo para digitar CPF
    campo_consulta = driver.find_element(By.XPATH, "//input[@id='cpf']")
    sleep(1)
    #Apaga o campo antes de enviar novamente
    campo_consulta.clear()

    campo_consulta.click()
    sleep(1)

    #Envia o CPF a ser digitado
    campo_consulta.send_keys(cpf)
    
    # Espera até que o botão de consulta esteja clicável
    # botao_consulta = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='botaoValidarCPF']")))
# Poderia ser assim também...mas o tempo seria cronometrado e teria que existir o "from time import sleep"
    sleep(1)
    #Identifica o botão de consulta
    botao_consulta = driver.find_element(By.XPATH, "//button[@id='botaoValidarCPF']")
    sleep(1)
    #Clica no botão
    botao_consulta.click()

      # Espera até que o resultado da validação esteja disponível no textarea
    # validacao = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='alert alert-danger']")))
# Poderia ser assim também...mas o tempo seria cronometrado e teria que existir o "from time import sleep"
    sleep(1)
    validacao = driver.find_element(By.XPATH, "//div[@role='alert']")

    if validacao.text == 'CPF Válido!': 
        #Para testar qual o índice da informação em um texto que vc quer,
        #basta testar print('texto-grande-muito.spli()) e vc terá o índice da palavra que vc quer
        status_validacao = 'Válido'
        validacao_CPF = openpyxl.load_workbook('cpf_clientes.xlsx')
        pagina_validacao = validacao_CPF['Sheet1']

        pagina_CPF.append([nome, cpf, status_validacao])

    else:
        status_validacao = 'Inválido!'
        validacao_CPF = openpyxl.load_workbook('cpf_clientes.xlsx')
        pagina_validacao = validacao_CPF['Sheet1']

        pagina_CPF.append([nome, cpf, status_validacao])

planilha_CPF.save('cpf_clientes.xlsx')


#Fecha após acabar a rotina
driver.quit()