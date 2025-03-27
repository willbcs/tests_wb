import openpyxl  # Biblioteca para abrir Excel
from selenium import webdriver  # Biblioteca para acessar a internet
from selenium.webdriver.common.by import By  # Permite encontrar elementos na página
from time import sleep
import os  # Para verificar se o arquivo existe

# Abrir planilha original
planilha_CPF = openpyxl.load_workbook('cpf_clientes.xlsx')
pagina_CPF = planilha_CPF['Sheet1']

# Verifica se a planilha de validação já existe
if os.path.exists('validacao.xlsx'):
    validacao_CPF = openpyxl.load_workbook('validacao.xlsx')
    pagina_validacao = validacao_CPF.active
else:
    validacao_CPF = openpyxl.Workbook()
    pagina_validacao = validacao_CPF.active
    pagina_validacao.append(["Nome", "CPF", "Status"])  # Criando cabeçalhos

# Acessar o site e entrar com os CPFs
driver = webdriver.Chrome()
driver.get('https://www.geradorcpf.com/validar-cpf.htm')

# Ler os CPFs a partir da linha 2
for linha in pagina_CPF.iter_rows(min_row=2, values_only=True):
    nome, cpf = linha

    sleep(4)

    # Localiza o campo para digitar CPF
    campo_consulta = driver.find_element(By.XPATH, "//input[@id='cpf']")
    sleep(1)
    campo_consulta.clear()
    campo_consulta.click()
    sleep(1)
    campo_consulta.send_keys(cpf)

    sleep(1)

    # Identifica e clica no botão de consulta
    botao_consulta = driver.find_element(By.XPATH, "//button[@id='botaoValidarCPF']")
    sleep(1)
    botao_consulta.click()

    sleep(1)

    # Captura o resultado da validação
    validacao = driver.find_element(By.XPATH, "//div[@role='alert']")
    
    status_validacao = "Válido" if "CPF Válido!" in validacao.text else "Inválido"

    # Adiciona os dados à planilha de validação
    pagina_validacao.append([nome, cpf, status_validacao])

# Salva os resultados na planilha `validacao.xlsx`
validacao_CPF.save('validacao_CPF.xlsx')

# Fecha após finalizar a rotina
driver.quit()
